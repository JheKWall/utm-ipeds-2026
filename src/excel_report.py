"""Build the Excel deliverable from the query results and charts.

Produces output/UTM_IPEDS_Analysis.xlsx with a documentation sheet, one sheet per
query, and the three charts embedded. Formatting (number formats, frozen panes,
column widths) is applied here; final polish is done by hand in Excel.
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)

# query file -> (sheet name, chart to embed, {column: number format})
SHEETS = {
    "q1": ("Outcomes", "1_outcomes.png", {
        "completion_rate": "0.0%", "still_enrolled_rate": "0.0%",
        "transfer_out_rate": "0.0%", "unknown_outcome_rate": "0.0%",
        "completion_or_transfer_rate": "0.0%",
        "completion_pct": "0.0", "completion_or_transfer_pct": "0.0",
    }),
    "q2": ("Cost series", "2_cost_of_living.png", {
        "food_housing": "$#,##0", "other_expenses": "$#,##0",
        "living_cost_total": "$#,##0", "books_supplies": "$#,##0",
        "tuition_fees_instate": "$#,##0",
    }),
    "q3": ("Combined", "3_combined.png", {
        "living_cost_2015_16": "$#,##0", "living_cost_2023_24": "$#,##0",
        "tuition_2015_16": "$#,##0", "tuition_2023_24": "$#,##0",
        "living_cost_change_pct": "0.0",
        "completion_rate": "0.0%", "still_enrolled_rate": "0.0%",
        "transfer_out_rate": "0.0%", "unknown_outcome_rate": "0.0%",
        "completion_or_transfer_rate": "0.0%",
    }),
    "q4": ("Window check", None, {
        "completion_pct_6yr": "0.0", "completion_pct_8yr": "0.0",
        "gain_6_to_8_pct": "0.0", "still_enrolled_pct_at_8yr": "0.0",
    }),
    "q5": ("Cohort trend", "5_cohort_trend.png", {
        "completion_pct": "0.0", "completion_or_transfer_pct": "0.0",
    }),
    "q6": ("Cost at entry", "6_within_institution.png", {
        "completion_pct": "0.0", "food_housing": "$#,##0",
        "other_expenses": "$#,##0", "living_cost_at_entry": "$#,##0",
        "tuition_at_entry": "$#,##0",
    }),
}

DOC_LINES = [
    ("UTM IPEDS Summer 2026 — Student Completion Outcomes", "title"),
    ("", ""),
    ("Research questions", "head"),
    ("1. Among selected institutions near UT Martin, what share of entering "
     "first-time, full-time freshmen complete their program?", ""),
    ("2. Is there a relationship between local cost of living and that completion "
     "rate?", ""),
    ("", ""),
    ("Data sources", "head"),
    ("IPEDS HD2025 — institution directory (names, sector, county)", ""),
    ("IPEDS OM2023 (final/revised) — Outcome Measures, first-time full-time cohort", ""),
    ("IPEDS IC2016_AY, IC2019_AY, IC2023_AY — cost of attendance, 2013-14 to 2023-24", ""),
    ("Downloaded from https://nces.ed.gov/ipeds/complete-data-files/", ""),
    ("", ""),
    ("Population", "head"),
    ("First-time full-time entering cohorts (OMCHRT = 10) for eight entry years, "
     "2009-10 through 2016-17, each measured at its own eight-year status point. "
     "A cohort comprises all degree/certificate-seeking undergraduates entering "
     "across a full 12-month period (July 1 - June 30), so spring and summer "
     "entrants are included.", ""),
    ("The 2015-16 cohort (measured Aug 31, 2023) is used for the headline "
     "single-cohort figures; all eight are used for the trend.", ""),
    ("", ""),
    ("What the cost figures are", "head"),
    ("Labelled 'cost of attendance: living expenses' throughout, using IPEDS's own "
     "terminology. This is the institution's ESTIMATED student budget for someone "
     "living off campus and not with family — food and housing plus other expenses "
     "(transportation, laundry, personal). It EXCLUDES tuition and fees, which are "
     "reported separately.", ""),
    ("It is an administrative estimate produced by a financial aid office, not a "
     "measurement of local prices. Institutions use differing methodologies, and "
     "the figures cap federal aid eligibility. They track local costs and are a "
     "reasonable proxy, but they are not a general cost-of-living index for the "
     "area — BEA Regional Price Parities would be that, and were evaluated and "
     "rejected because per-state nonmetropolitan values do not exist.", ""),
    ("", ""),
    ("Measures", "head"),
    ("Completion rate — received an award, divided by cohort", ""),
    ("Still-enrolled rate — no award, still enrolled at the reporting institution", ""),
    ("Transfer-out rate — no award, subsequently enrolled elsewhere", ""),
    ("Unknown-outcome rate — no award, no subsequent enrollment found", ""),
    ("Completion-or-transfer rate — awarded plus enrolled elsewhere, over cohort", ""),
    ("", ""),
    ("Important limitations", "head"),
    ("• 'Unknown outcome' is NOT a dropout rate. It is a residual that also absorbs "
     "students who enrolled somewhere the National Student Clearinghouse does not "
     "cover.", ""),
    ("• Five institutions cannot support a correlation. Using 2015-16 cost data the "
     "correlation with completion is -0.55; using 2023-24 data it is +0.82. Same "
     "institutions, same completion rates, opposite conclusions — which is why no "
     "trend line is fitted and no inferential statistic is reported.", ""),
    ("• Three of the five are two-year colleges. Program length, student "
     "demographics, and institutional mission all differ at once and cannot be "
     "separated at this sample size. Higher university completion is NOT evidence "
     "of institutional quality.", ""),
    ("• Transferring out is mission success for a community college, not failure, "
     "which is why completion-or-transfer is reported alongside completion.", ""),
    ("• IPEDS records what happened to students, never why. No causal claim about "
     "reasons for leaving is supportable.", ""),
    ("• Institutions self-report cost estimates and some do not update them "
     "annually. Jackson State reports a near-flat figure across all eleven years; "
     "Arkansas Northeastern's 2013-14 value falls 43% the following year and is a "
     "suspected reporting error.", ""),
    ("• Cost data covers 2013-14 to 2023-24; the cohort entered in 2015-16. Figures "
     "for later years describe the locality, not these students' own expenses.", ""),
]


def _write_doc_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 110
    for i, (text, kind) in enumerate(DOC_LINES, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "title":
            cell.font = TITLE_FONT
        elif kind == "head":
            cell.font = Font(bold=True, size=12, color="1F3864")
        if len(text) > 100:
            ws.row_dimensions[i].height = 30


def _write_data_sheet(ws, df: pd.DataFrame, formats: dict) -> None:
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for record in df.itertuples(index=False):
        ws.append([None if pd.isna(v) else v for v in record])

    for idx, column in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        longest = max([len(str(column))] + [len(str(v)) for v in df[column].head(50)])
        ws.column_dimensions[letter].width = min(max(longest + 2, 11), 38)
        if column in formats:
            for cell in ws[letter][1:]:
                cell.number_format = formats[column]

    ws.freeze_panes = "A2"


def main() -> None:
    wb = Workbook()
    _write_doc_sheet(wb.active)
    wb.active.title = "Documentation"

    for key, (sheet_name, chart, formats) in SHEETS.items():
        path = config.QUERIES / f"{key}.csv"
        if not path.exists():
            print(f"  skipping {key}: {path.name} not found")
            continue
        df = pd.read_csv(path)
        ws = wb.create_sheet(sheet_name)
        _write_data_sheet(ws, df, formats)

        if chart:
            image_path = config.CHARTS / chart
            if image_path.exists():
                anchor_row = len(df) + 4
                img = XLImage(str(image_path))
                # Scale to a readable width without distorting aspect ratio.
                scale = 900 / img.width
                img.width, img.height = int(img.width * scale), int(img.height * scale)
                ws.add_image(img, f"A{anchor_row}")
        print(f"  {sheet_name}: {len(df)} rows")

    out = config.OUTPUT / "UTM_IPEDS_Analysis.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
